import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from pathlib import Path

from utils import (
    build_asf_output_stream,
    get_mapping_by_name,
    load_constant_values,
    load_all_mappings,
    load_override_mapping,
    save_mapping,
    suggest_mappings,
    write_loan_data_to_asf,
)

st.set_page_config(page_title="ASF Loan Tape Mapper", layout="wide")


def initialize_session_state():
    default_override_path = Path(__file__).with_name("mapping_overrides.yaml")
    default_constant_path = Path(__file__).with_name("constant_values.yaml")
    default_overrides = {}
    default_override_error = None
    default_constants = {}
    default_constant_error = None

    if default_override_path.exists():
        try:
            default_overrides = load_override_mapping(default_override_path)
        except Exception as exc:  # pylint: disable=broad-except
            default_override_error = str(exc)

    if default_constant_path.exists():
        try:
            default_constants = load_constant_values(default_constant_path)
        except Exception as exc:  # pylint: disable=broad-except
            default_constant_error = str(exc)

    defaults = {
        "field_mappings": {},
        "config": {},
        "last_threshold": None,
        "override_mapping": default_overrides,
        "default_override_error": default_override_error,
        "constant_values": default_constants,
        "constant_values_error": default_constant_error,
    }

    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

    if "field_mappings" not in st.session_state:
        st.session_state["field_mappings"] = {}


def load_asf_template(template_file, sheet_name=None):
    """
    Load ASF template workbook and return (workbook, worksheet).
    If sheet_name is None, use the first sheet.
    """

    template_file.seek(0)
    workbook = load_workbook(template_file)

    worksheet_name = sheet_name or workbook.sheetnames[0]
    worksheet = workbook[worksheet_name]

    return workbook, worksheet


def get_asf_fields(ws, header_row=1):
    """
    Read the header row from the ASF worksheet and return a list of non-empty field names.
    """

    rows = ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    header_cells = next(rows, [])

    return [str(cell).strip() for cell in header_cells if cell is not None and str(cell).strip()]


def load_tape_into_dataframe(file):
    filename = file.name.lower()

    if filename.endswith(".csv"):
        return pd.read_csv(file)

    if filename.endswith(".xlsx"):
        return pd.read_excel(file)

    raise ValueError("Unsupported file type. Please upload a .csv or .xlsx file.")


def render_mapping_editor(
    asf_fields,
    tape_fields,
    current_mapping,
    file_key_prefix,
    threshold,
    constant_values=None,
    tape_samples=None,
):
    """
    Render a side-by-side editor:
    - Left: ASF field (text)
    - Middle: selectbox over ["(unmapped)"] + tape_fields
    - Right: match score (read-only)
    - Far Right: preview of the first few values from the selected tape field
    Return updated mapping dict in same structure as current_mapping.
    Use file_key_prefix + asf_field for Streamlit widget keys.
    """

    updated_mapping = {}
    tape_samples = tape_samples or {}
    used_source_fields = {
        info.get("source_field")
        for info in current_mapping.values()
        if info.get("source_field") not in (None, "(unmapped)")
        and not info.get("use_constant", False)
    }

    header_cols = st.columns([3, 3, 1, 2], gap="small")
    with header_cols[0]:
        st.caption("ASF Field")
    with header_cols[1]:
        st.caption("Source Field")
    with header_cols[2]:
        st.caption("Match Score")
    with header_cols[3]:
        st.caption("Preview of first row")

    for asf_field in asf_fields:
        source_field = None
        score = None
        use_constant = False

        if asf_field in current_mapping:
            source_field = current_mapping[asf_field].get("source_field")
            score = current_mapping[asf_field].get("score")
            use_constant = current_mapping[asf_field].get("use_constant", False)

        col1, col2, col3, col4 = st.columns([3, 3, 1, 2])

        with col1:
            st.text(asf_field)

        widget_key = f"{file_key_prefix}{asf_field}_{threshold}"
        widget_value = st.session_state.get(widget_key)

        selection_options = ["(unmapped)"]
        constant_label = None
        if constant_values and asf_field in constant_values:
            constant_label = f"(constant: {constant_values[asf_field]})"
            selection_options.append(constant_label)

        available_fields = []
        for field in tape_fields:
            if field == source_field or field == widget_value or field not in used_source_fields:
                available_fields.append(field)

        for field in available_fields:
            if field not in selection_options:
                selection_options.append(field)

        default_value = "(unmapped)"
        for candidate in (widget_value, source_field, constant_label, "(unmapped)"):
            if candidate in selection_options:
                default_value = candidate
                break

        with col2:
            selected_source = st.selectbox(
                "Source Field",
                options=selection_options,
                index=selection_options.index(default_value),
                key=widget_key,
                label_visibility="collapsed",
            )

        with col3:
            st.text(str(score))

        sample_text = ""
        if selected_source in tape_samples:
            sample_text = tape_samples[selected_source]
        elif selected_source == constant_label or (use_constant and constant_label):
            sample_text = f"Constant: {constant_values.get(asf_field)}"

        with col4:
            st.caption(sample_text if sample_text else " ")

        final_source = None if selected_source in ("(unmapped)", constant_label) else selected_source
        updated_mapping[asf_field] = {
            "source_field": final_source,
            "score": score,
            "use_constant": selected_source == constant_label,
        }

    return updated_mapping


def render_mapping_persistence_controls(
    tape_file_name,
    tape_columns,
    asf_fields,
    threshold,
):
    st.markdown("#### Save & Load Mapping")

    try:
        saved_mappings = load_all_mappings()
    except ValueError as exc:  # pylint: disable=broad-except
        st.warning(str(exc))
        saved_mappings = {}

    mapping_name = st.text_input(
        "Mapping name/label",
        key=f"mapping_label_{tape_file_name}",
        help="Provide a name to save or overwrite a mapping",
    )

    overwrite = False
    if mapping_name and mapping_name.strip() in saved_mappings:
        overwrite = st.checkbox(
            "Overwrite existing mapping with this name",
            key=f"overwrite_{tape_file_name}",
        )
        st.caption("A mapping with this name already exists.")

    if st.button("Save mapping", key=f"save_mapping_{tape_file_name}"):
        mapping = st.session_state["field_mappings"].get(tape_file_name, {})
        if not mapping_name or not mapping_name.strip():
            st.error("Please provide a mapping name before saving.")
        elif not mapping:
            st.error("No mapping is available to save.")
        elif not any(
            info.get("source_field")
            or info.get("use_constant")
            for info in mapping.values()
        ):
            st.error("Define at least one mapped or constant field before saving.")
        else:
            try:
                save_mapping(
                    mapping_name.strip(),
                    mapping,
                    tape_columns=tape_columns,
                    overwrite=overwrite,
                )
            except ValueError as exc:  # pylint: disable=broad-except
                st.error(str(exc))
            except OSError as exc:  # pylint: disable=broad-except
                st.error(f"Could not save mapping: {exc}")
            else:
                if overwrite:
                    st.success(f"Mapping '{mapping_name}' overwritten.")
                else:
                    st.success(f"Mapping '{mapping_name}' saved.")

    st.divider()

    mapping_names = sorted(saved_mappings.keys())
    selected_mapping = st.selectbox(
        "Load saved mapping",
        options=[""] + mapping_names,
        format_func=lambda x: "Select a mapping" if x == "" else x,
        key=f"load_mapping_select_{tape_file_name}",
    )

    if st.button("Load mapping", key=f"load_mapping_{tape_file_name}"):
        if not selected_mapping:
            st.warning("Choose a mapping to load.")
            return

        try:
            saved_mapping = get_mapping_by_name(selected_mapping)
        except ValueError as exc:  # pylint: disable=broad-except
            st.error(str(exc))
            return
        except OSError as exc:  # pylint: disable=broad-except
            st.error(f"Could not read saved mappings: {exc}")
            return

        if not saved_mapping:
            st.error(f"Mapping '{selected_mapping}' was not found.")
            return

        mapping_data = saved_mapping.get("mapping", {})
        updated_mapping = st.session_state["field_mappings"].get(
            tape_file_name, {}
        )

        constant_values = st.session_state.get("constant_values") or {}
        missing_columns = []
        for asf_field in asf_fields:
            if asf_field not in mapping_data:
                continue

            new_info = dict(mapping_data.get(asf_field) or {})
            source_field = new_info.get("source_field")
            if source_field and source_field not in tape_columns:
                missing_columns.append(source_field)
                new_info["source_field"] = None
                new_info["score"] = None

            updated_mapping[asf_field] = new_info

            widget_key = f"{tape_file_name}_{asf_field}_{threshold}"
            widget_value = new_info.get("source_field")
            if new_info.get("use_constant") and asf_field in constant_values:
                widget_value = f"(constant: {constant_values[asf_field]})"
            if widget_value is None:
                st.session_state.pop(widget_key, None)
            else:
                st.session_state[widget_key] = widget_value

        st.session_state["field_mappings"][tape_file_name] = updated_mapping

        if missing_columns:
            missing_list = ", ".join(sorted(set(missing_columns)))
            st.warning(
                "Mapping loaded, but these source columns were not found in the "
                f"current file: {missing_list}. They have been left unmapped."
            )

        st.success(f"Mapping '{selected_mapping}' loaded.")


def render_sidebar():
    st.sidebar.header("Setup")
    st.sidebar.subheader("ASF template upload")
    asf_template_file = st.sidebar.file_uploader(
        "Upload ASF template",
        type=["xlsx"],
        key="asf_template_file",
    )

    st.sidebar.subheader("Loan tape upload")
    tape_files = st.sidebar.file_uploader(
        "Upload loan tape files",
        type=["xlsx", "csv"],
        accept_multiple_files=True,
        key="tape_files",
    )

    threshold = st.sidebar.slider(
        "Fuzzy match threshold", min_value=0, max_value=100, value=80
    )

    st.sidebar.caption(
        "Mapping suggestions use the bundled mapping_overrides.yaml. "
        "Edit that file to adjust alias preferences."
    )

    if st.session_state.get("default_override_error"):
        st.sidebar.warning(
            "Default mapping_overrides.yaml could not be loaded: "
            f"{st.session_state['default_override_error']}. "
            "Override suggestions will fallback to exact field names."
        )
    else:
        st.sidebar.caption(
            f"Loaded {len(st.session_state.get('override_mapping', {}))} override entries"
        )

    st.sidebar.caption(
        f"Loaded {len(st.session_state.get('constant_values', {}))} constant field values "
        "from constant_values.yaml (if present)."
    )
    if st.session_state.get("constant_values_error"):
        st.sidebar.warning(
            "constant_values.yaml could not be loaded: "
            f"{st.session_state['constant_values_error']}"
        )

    generate_clicked = st.sidebar.button("Generate ASF Files")

    return asf_template_file, tape_files, threshold, False, generate_clicked


def render_main_content(asf_template_file, tape_files, threshold, override_changed, sidebar_generate_clicked):
    st.title("ASF Loan Tape Mapper")
    st.subheader("Workflow")
    st.markdown(
        """
        1. Upload ASF template
        2. Upload loan tapes
        3. Review mapping
        4. Download ASF output
        """
    )

    if not asf_template_file and not tape_files:
        st.info("Upload ASF template and loan tape files to begin.")
        return

    asf_fields = []
    previous_threshold = st.session_state.get("last_threshold")
    threshold_changed = previous_threshold != threshold
    st.session_state["last_threshold"] = threshold

    if asf_template_file:
        st.markdown(f"**ASF template uploaded:** {asf_template_file.name}")
        st.caption("ASF template uploaded")

        wb, ws = load_asf_template(asf_template_file)
        asf_fields = get_asf_fields(ws)

    if threshold_changed or override_changed:
        # Clear stored mappings and widget state so new threshold suggestions become defaults
        for tape_name, mapping in list(st.session_state["field_mappings"].items()):
            for asf_field in mapping.keys():
                st.session_state.pop(
                    f"{tape_name}_{asf_field}_{previous_threshold}", None
                )
        st.session_state["field_mappings"] = {}

    if tape_files:
        tab_labels = [tape_file.name for tape_file in tape_files]
        tabs = st.tabs(tab_labels)

        for tape_file, tab in zip(tape_files, tabs):
            with tab:
                st.markdown(f"### {tape_file.name}")
                dataframe = load_tape_into_dataframe(tape_file)
                st.dataframe(dataframe.head(), use_container_width=True)

                tape_cols = list(dataframe.columns)
                sample_preview = {}
                if not dataframe.empty:
                    for col in tape_cols:
                        val = dataframe[col].iloc[0]
                        sample_preview[col] = "" if pd.isna(val) else str(val)

                if threshold_changed or tape_file.name not in st.session_state["field_mappings"]:
                    mapping_suggestions = suggest_mappings(
                        asf_fields,
                        tape_cols,
                        threshold,
                        overrides=st.session_state.get("override_mapping"),
                    )

                    const_vals = st.session_state.get("constant_values") or {}
                    for asf_field, const_val in const_vals.items():
                        if asf_field not in asf_fields:
                            continue
                        mapping_suggestions.setdefault(
                            asf_field, {"source_field": None, "score": None}
                        )
                        mapping_suggestions[asf_field]["use_constant"] = True

                    st.session_state["field_mappings"][tape_file.name] = (
                        mapping_suggestions
                    )

                mapping_dict = st.session_state["field_mappings"][tape_file.name]
                ordered_fields = [field for field in asf_fields if field in mapping_dict]
                chunk_size = (len(ordered_fields) + 2) // 3
                first_fields = ordered_fields[:chunk_size]
                second_fields = ordered_fields[chunk_size : 2 * chunk_size]
                third_fields = ordered_fields[2 * chunk_size :]

                st.markdown("#### Field Mappings")
                col1, col2, col3 = st.columns(3)

                with col1:
                    updated_first = render_mapping_editor(
                        first_fields,
                        tape_cols,
                        mapping_dict,
                        file_key_prefix=f"{tape_file.name}_",
                        threshold=threshold,
                        constant_values=st.session_state.get("constant_values"),
                        tape_samples=sample_preview,
                    )

                with col2:
                    updated_second = render_mapping_editor(
                        second_fields,
                        tape_cols,
                        mapping_dict,
                        file_key_prefix=f"{tape_file.name}_",
                        threshold=threshold,
                        constant_values=st.session_state.get("constant_values"),
                        tape_samples=sample_preview,
                    )

                with col3:
                    updated_third = render_mapping_editor(
                        third_fields,
                        tape_cols,
                        mapping_dict,
                        file_key_prefix=f"{tape_file.name}_",
                        threshold=threshold,
                        constant_values=st.session_state.get("constant_values"),
                        tape_samples=sample_preview,
                    )

                updated_mapping = dict(mapping_dict)
                updated_mapping.update(updated_first)
                updated_mapping.update(updated_second)
                updated_mapping.update(updated_third)
                st.session_state["field_mappings"][tape_file.name] = updated_mapping

                render_mapping_persistence_controls(
                    tape_file_name=tape_file.name,
                    tape_columns=tape_cols,
                    asf_fields=asf_fields,
                    threshold=threshold,
                )

        if asf_template_file:
            generate_clicked = sidebar_generate_clicked or st.button("Generate ASF Files")
            if generate_clicked:
                for tape_file in tape_files:
                    df = load_tape_into_dataframe(tape_file)
                    mapping = st.session_state["field_mappings"].get(
                        tape_file.name, {}
                    )
                    wb, ws = load_asf_template(asf_template_file)
                    asf_fields = get_asf_fields(ws)
                    write_loan_data_to_asf(
                        ws,
                        start_row=2,
                        asf_fields=asf_fields,
                        df=df,
                        mapping=mapping,
                        constant_values=st.session_state.get("constant_values"),
                    )
                    output_stream = build_asf_output_stream(wb)
                    st.download_button(
                        label=f"Download ASF-mapped file for {tape_file.name}",
                        data=output_stream,
                        file_name=f"ASF_{tape_file.name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )


def main():
    initialize_session_state()
    asf_template_file, tape_files, threshold, override_changed, sidebar_generate_clicked = render_sidebar()
    render_main_content(asf_template_file, tape_files, threshold, override_changed, sidebar_generate_clicked)


if __name__ == "__main__":
    main()

# To do
# Only display matching score for mapped fields
# Add "unmapped" option to selectboxes
# Only display matching score for mapped fields
# Add "unmapped" option to selectboxes
